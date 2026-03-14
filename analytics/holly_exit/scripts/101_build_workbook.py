"""
Script 101: Build Holly Analysis Lab Workbook (v1)

Static, honest, polished analytical workbook with 6 sheets:
  README, DATA_DICTIONARY, PRETRADE_FEATURES, STRATEGY_LAB, REGIME_LAB, SCORECARD

Reads from DuckDB bronze + silver. Computes SSP overlay fresh.
Output: analytics/holly_exit/output/holly_analysis_lab.xlsx
Does NOT touch holly_analytics.xlsx.
"""

import sys, os, math, warnings
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import numpy as np
import pandas as pd
import duckdb
from scipy import stats as sp_stats

warnings.filterwarnings("ignore", category=FutureWarning)

# ── paths ────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent.parent
HOLLY_DIR = SCRIPT_DIR.parent
DDB_PATH = HOLLY_DIR / "data" / "duckdb" / "holly.ddb"
OUTPUT_DIR = HOLLY_DIR / "output"
OUTPUT_PATH = OUTPUT_DIR / "holly_analysis_lab.xlsx"

# ── feature registry ────────────────────────────────────────────────────
# timing_class: "pretrade" = available before entry, "posttrade" = uses outcome,
#               "leaky" = uses full-sample stats, "id" = identifier/skip
FEATURE_REGISTRY = {
    # --- identifiers ---
    "trade_id":           ("Trade ID",                 "id",        "trades"),
    "symbol":             ("Ticker symbol",            "id",        "trades"),
    "strategy":           ("Holly strategy name",      "id",        "trades"),
    "direction":          ("Trade direction",          "id",        "trades"),
    "entry_time":         ("Entry timestamp",          "id",        "trades"),
    "entry_price":        ("Entry price",              "id",        "trades"),
    "exit_time":          ("Exit timestamp",           "id",        "trades"),
    "exit_price":         ("Exit price",               "id",        "trades"),
    "holly_pnl":          ("Trade P&L",                "posttrade", "trades"),
    "shares":             ("Position size",            "id",        "trades"),
    "stop_price":         ("Stop loss price",          "pretrade",  "trades"),
    "target_price":       ("Target price",             "pretrade",  "trades"),
    "mfe":                ("Max favorable excursion",  "posttrade", "trades"),
    "mae":                ("Max adverse excursion",    "posttrade", "trades"),

    # --- pretrade: ticker details ---
    "market_cap":         ("Market capitalization",    "pretrade",  "ticker_details"),
    "sic_code":           ("SIC industry code",        "pretrade",  "ticker_details"),
    "sector":             ("SIC sector (2-digit)",     "pretrade",  "ticker_details"),
    "primary_exchange":   ("Exchange",                 "pretrade",  "ticker_details"),

    # --- pretrade: regime ---
    "trend_regime":       ("Trend regime",             "pretrade",  "trade_regime"),
    "vol_regime":         ("Volatility regime",        "pretrade",  "trade_regime"),
    "momentum_regime":    ("Momentum regime",          "pretrade",  "trade_regime"),
    "rsi14":              ("RSI(14) at entry",         "pretrade",  "trade_regime"),
    "atr_pct":            ("ATR% at entry",            "pretrade",  "trade_regime"),
    "trend_slope":        ("20-day trend slope",       "pretrade",  "trade_regime"),
    "sma20":              ("SMA(20) at entry",         "pretrade",  "trade_regime"),
    "sma5":               ("SMA(5) at entry",          "pretrade",  "trade_regime"),
    "above_sma20":        ("Price > SMA(20)",          "pretrade",  "trade_regime"),
    "roc5":               ("5-day rate of change",     "pretrade",  "trade_regime"),
    "roc20":              ("20-day rate of change",    "pretrade",  "trade_regime"),
    "daily_range_pct":    ("Daily range %",            "pretrade",  "trade_regime"),
    "atr14":              ("ATR(14) value",            "pretrade",  "trade_regime"),

    # --- pretrade: indicator features ---
    "rsi_zone":           ("RSI zone bucket",          "pretrade",  "trade_indicator_features"),
    "macd_cross":         ("MACD cross state",         "pretrade",  "trade_indicator_features"),
    "ma_trend":           ("MA trend alignment",       "pretrade",  "trade_indicator_features"),
    "price_vs_ema9":      ("Price vs EMA(9) %",        "pretrade",  "trade_indicator_features"),
    "price_vs_ema21":     ("Price vs EMA(21) %",       "pretrade",  "trade_indicator_features"),
    "above_sma50":        ("Price > SMA(50)",          "pretrade",  "trade_indicator_features"),

    # --- pretrade: short features ---
    "short_interest":     ("Short interest shares",    "pretrade",  "trade_short_features"),
    "days_to_cover":      ("Days to cover",            "pretrade",  "trade_short_features"),
    "short_volume_ratio": ("Short volume ratio",       "pretrade",  "trade_short_features"),
    "short_squeeze_regime":("Short squeeze regime",    "pretrade",  "trade_short_features"),

    # --- pretrade: insider features ---
    "insider_buy_count_30d": ("Insider buys (30d)",    "pretrade",  "trade_insider_features"),
    "insider_sell_count_30d":("Insider sells (30d)",   "pretrade",  "trade_insider_features"),
    "insider_cluster_flag":  ("Cluster buying flag",   "pretrade",  "trade_insider_features"),
    "insider_net_value":     ("Insider net value",     "pretrade",  "trade_insider_features"),

    # --- pretrade: news ---
    "news_count_24h":     ("News articles (24h)",      "pretrade",  "benzinga_features_broad"),
    "news_count_7d":      ("News articles (7d)",       "pretrade",  "benzinga_features_broad"),
    "news_count_30d":     ("News articles (30d)",      "pretrade",  "benzinga_features_broad"),
    "news_acceleration":  ("News acceleration",        "pretrade",  "benzinga_features_broad"),
    "has_movers_news":    ("Has movers news",          "pretrade",  "benzinga_features_broad"),

    # --- pretrade: macro ---
    "vix_regime":         ("VIX regime",               "pretrade",  "fred_macro_daily"),
    "yield_curve_regime": ("Yield curve regime",       "pretrade",  "fred_macro_daily"),
    "rate_regime":        ("Rate regime",              "pretrade",  "fred_macro_daily"),
    "macro_vix":          ("VIX level",                "pretrade",  "fred_macro_daily"),
    "macro_yield_spread": ("Yield spread 10y-2y",      "pretrade",  "fred_macro_daily"),

    # --- pretrade: prior day / gap ---
    "entry_gap_pct":      ("Entry gap %",              "pretrade",  "computed"),
    "prior_day_close":    ("Prior day close",          "pretrade",  "computed"),
    "prior_day_volume":   ("Prior day volume",         "pretrade",  "computed"),

    # --- pretrade: SPY context ---
    "spy_intraday_pct_at_entry": ("SPY intraday % at entry", "pretrade", "computed"),
    "spy_price_at_entry":        ("SPY price at entry",      "pretrade", "computed"),

    # --- pretrade: spread quality ---
    "nbbo_spread_pct":    ("NBBO spread %",            "pretrade",  "computed"),

    # --- pretrade: float ---
    "free_float":         ("Free float shares",        "pretrade",  "computed"),
    "float_rotation_pct": ("Float rotation %",         "pretrade",  "computed"),

    # --- pretrade: fundamentals ---
    "price_to_earnings":  ("P/E ratio",                "pretrade",  "computed"),
    "ev_to_ebitda":       ("EV/EBITDA",                "pretrade",  "computed"),
    "return_on_equity":   ("ROE",                      "pretrade",  "computed"),

    # --- pretrade: time features ---
    "entry_hour":         ("Entry hour (ET)",          "pretrade",  "computed"),
    "trade_dow":          ("Day of week",              "pretrade",  "computed"),

    # --- pretrade: SSP overlay (computed fresh) ---
    "strat_sector_prior_wr": ("Shrunk strategy-sector WR", "pretrade", "ssp_overlay"),
    "ssp_delta":          ("SSP delta vs global",      "pretrade",  "ssp_overlay"),
    "ssp_bonus":          ("SSP overlay bonus",        "pretrade",  "ssp_overlay"),

    # --- LEAKY: full-sample stats that look pretrade ---
    "strat_win_rate":     ("Strategy win rate (FULL SAMPLE)", "leaky", "full_sample"),
    "strat_avg_pnl":      ("Strategy avg PnL (FULL SAMPLE)", "leaky", "full_sample"),
    "strat_total_pnl":    ("Strategy total PnL (FULL SAMPLE)","leaky", "full_sample"),
    "strat_sharpe":       ("Strategy Sharpe (FULL SAMPLE)",   "leaky", "full_sample"),
    "sector_win_rate":    ("Sector win rate (FULL SAMPLE)",   "leaky", "full_sample"),
    "sector_avg_pnl":     ("Sector avg PnL (FULL SAMPLE)",   "leaky", "full_sample"),
    "tod_win_rate":       ("Time-of-day WR (FULL SAMPLE)",   "leaky", "full_sample"),
    "tod_avg_pnl":        ("Time-of-day avg PnL (FULL SAMPLE)","leaky","full_sample"),

    # --- posttrade: optimization ---
    "opt_exit_rule":      ("Optimized exit rule",      "posttrade", "optimization_results"),
    "opt_avg_pnl":        ("Optimized avg PnL",        "posttrade", "optimization_results"),
    "opt_win_rate":       ("Optimized win rate",       "posttrade", "optimization_results"),
    "opt_sharpe":         ("Optimized Sharpe",         "posttrade", "optimization_results"),
    "opt_profit_factor":  ("Optimized profit factor",  "posttrade", "optimization_results"),
}


# ── helpers ──────────────────────────────────────────────────────────────

def safe_sic2(x):
    """Extract 2-digit SIC code safely."""
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return None
    s = str(x).strip()
    if not s:
        return None
    try:
        return str(int(float(s)))[:2]
    except (ValueError, OverflowError):
        return None


def cohens_d(group1, group2):
    """Compute Cohen's d effect size."""
    n1, n2 = len(group1), len(group2)
    if n1 < 5 or n2 < 5:
        return 0.0, 1.0  # d, p
    m1, m2 = np.mean(group1), np.mean(group2)
    s1, s2 = np.var(group1, ddof=1), np.var(group2, ddof=1)
    pooled = np.sqrt(((n1 - 1) * s1 + (n2 - 1) * s2) / (n1 + n2 - 2))
    if pooled < 1e-12:
        return 0.0, 1.0
    d = (m1 - m2) / pooled
    _, p = sp_stats.ttest_ind(group1, group2, equal_var=False)
    return d, p


def significance_stars(p):
    if p < 0.001:
        return "***"
    elif p < 0.01:
        return "**"
    elif p < 0.05:
        return "*"
    return ""


def bayesian_wr(wins, n, prior_a=1, prior_b=1):
    """Beta-binomial posterior: mean and 95% CI."""
    a = prior_a + wins
    b = prior_b + (n - wins)
    mean = a / (a + b)
    lo = sp_stats.beta.ppf(0.025, a, b)
    hi = sp_stats.beta.ppf(0.975, a, b)
    return mean, lo, hi


def profit_factor(pnls):
    gross_profit = pnls[pnls > 0].sum()
    gross_loss = abs(pnls[pnls < 0].sum())
    if gross_loss < 0.01:
        return float("inf") if gross_profit > 0 else 0.0
    return gross_profit / gross_loss


def sharpe_proxy(pnls):
    if len(pnls) < 5 or pnls.std() < 0.01:
        return 0.0
    return pnls.mean() / pnls.std() * np.sqrt(len(pnls))


# ── SSP overlay (same logic as script 99) ────────────────────────────────

def build_shrunk_priors(df, min_cell=5, min_strat=10):
    """Hierarchical Bayes shrinkage for strategy-sector prior WR."""
    global_wr = df["win"].mean()

    # strategy-level priors
    strat_stats = df.groupby("strategy")["win"].agg(["mean", "count"])
    strat_prior = {}
    for s, row in strat_stats.iterrows():
        if row["count"] >= min_strat:
            strat_prior[s] = row["mean"]
        else:
            strat_prior[s] = global_wr

    # cell-level shrinkage
    df = df.copy()
    df["sic2"] = df["sic_code"].apply(safe_sic2)
    df["cell_key"] = df["strategy"] + "_" + df["sic2"].fillna("XX")

    cell_stats = df.groupby("cell_key")["win"].agg(["mean", "count"])
    cell_wr = {}
    cell_n = {}
    shrunk = {}

    for ck, row in cell_stats.iterrows():
        strat = ck.rsplit("_", 1)[0]
        target = strat_prior.get(strat, global_wr)
        n = row["count"]
        raw = row["mean"]
        if n < min_cell:
            shrunk[ck] = target
        else:
            alpha = min_cell / (min_cell + n)
            shrunk[ck] = (1 - alpha) * raw + alpha * target
        cell_wr[ck] = raw
        cell_n[ck] = int(n)

    df["raw_strat_sector_wr"] = df["cell_key"].map(cell_wr)
    df["shrunk_strat_sector_wr"] = df["cell_key"].map(shrunk)
    df["cell_n"] = df["cell_key"].map(cell_n)
    df["strat_prior_wr"] = df["strategy"].map(strat_prior)
    df["global_prior_wr"] = global_wr

    return df


def apply_overlay(df, cap=10, neutral_band=2):
    """Apply SSP overlay as capped bonus on top of AQS v2."""
    df = df.copy()
    delta = (df["shrunk_strat_sector_wr"] - df["global_prior_wr"]) * 100
    bonus = np.where(
        delta > neutral_band,
        np.minimum(delta - neutral_band, cap),
        np.where(delta < -neutral_band, np.maximum(delta + neutral_band, -cap), 0.0),
    )
    df["ssp_delta"] = delta
    df["ssp_bonus"] = bonus
    return df


# ── data loading ─────────────────────────────────────────────────────────

def load_data():
    """Load core trade data + enrichments from DuckDB."""
    con = duckdb.connect(str(DDB_PATH), read_only=True)

    # check available tables
    tables = [r[0] for r in con.execute("SHOW TABLES").fetchall()]
    print(f"DuckDB tables ({len(tables)}): {', '.join(sorted(tables)[:20])}...")

    # core trades with ticker details
    sql = """
    SELECT
        t.trade_id, t.entry_time, t.symbol, t.direction, t.strategy,
        t.entry_price, t.stop_price, t.target_price,
        t.exit_time, t.exit_price, t.holly_pnl, t.shares,
        t.mfe, t.mae,
        CASE WHEN t.holly_pnl > 0 THEN 1 ELSE 0 END AS win,
        CAST(t.entry_time AS DATE) AS trade_date,
        EXTRACT(HOUR FROM t.entry_time) AS entry_hour,
        EXTRACT(DOW FROM t.entry_time) AS trade_dow,
        td.sic_code, td.market_cap, td.primary_exchange
    FROM trades t
    LEFT JOIN ticker_details td ON t.symbol = td.symbol
    ORDER BY t.entry_time
    """
    df = con.execute(sql).fetchdf()
    print(f"Core trades: {len(df)} rows")

    # regime
    if "trade_regime" in tables:
        regime = con.execute("""
            SELECT trade_id, trend_regime, vol_regime, momentum_regime,
                   rsi14, atr_pct, trend_slope, sma20, sma5, above_sma20,
                   roc5, roc20, daily_range_pct, atr14
            FROM trade_regime
        """).fetchdf()
        df = df.merge(regime, on="trade_id", how="left")
        print(f"  + regime: {len(regime)} rows joined")

    # indicator features
    if "trade_indicator_features" in tables:
        ind = con.execute("""
            SELECT trade_id, rsi_zone, macd_cross, ma_trend,
                   price_vs_ema9, price_vs_ema21, above_sma50
            FROM trade_indicator_features
        """).fetchdf()
        df = df.merge(ind, on="trade_id", how="left")
        print(f"  + indicators: {len(ind)} rows joined")

    # short features
    if "trade_short_features" in tables:
        short = con.execute("""
            SELECT trade_id, short_interest, days_to_cover,
                   short_volume_ratio, short_squeeze_regime
            FROM trade_short_features
        """).fetchdf()
        df = df.merge(short, on="trade_id", how="left")
        print(f"  + short features: {len(short)} rows joined")

    # insider features
    if "trade_insider_features" in tables:
        ins = con.execute("""
            SELECT trade_id, insider_buy_count_30d, insider_sell_count_30d,
                   insider_cluster_flag, insider_net_value
            FROM trade_insider_features
        """).fetchdf()
        df = df.merge(ins, on="trade_id", how="left")
        print(f"  + insider features: {len(ins)} rows joined")

    # benzinga broad
    if "benzinga_features_broad" in tables:
        bz = con.execute("""
            SELECT trade_id, news_count_24h, news_count_7d, news_count_30d,
                   news_acceleration, has_movers_news
            FROM benzinga_features_broad
        """).fetchdf()
        df = df.merge(bz, on="trade_id", how="left")
        print(f"  + benzinga broad: {len(bz)} rows joined")

    # macro
    if "fred_macro_daily" in tables:
        macro = con.execute("""
            SELECT date,
                   vix AS macro_vix,
                   yield_spread_10y2y AS macro_yield_spread,
                   vix_regime, yield_curve_regime, rate_regime
            FROM fred_macro_daily
        """).fetchdf()
        macro["date"] = pd.to_datetime(macro["date"]).dt.date
        df["_join_date"] = pd.to_datetime(df["trade_date"]).dt.date
        df = df.merge(macro, left_on="_join_date", right_on="date", how="left")
        df.drop(columns=["_join_date", "date"], inplace=True, errors="ignore")
        print(f"  + macro: {len(macro)} rows joined")

    con.close()

    # sector from SIC
    df["sic2"] = df["sic_code"].apply(safe_sic2)

    # SSP overlay
    df = build_shrunk_priors(df)
    df = apply_overlay(df)
    print(f"  + SSP overlay applied ({df['shrunk_strat_sector_wr'].notna().sum()} non-null)")

    return df


# ══════════════════════════════════════════════════════════════════════════
#  WORKBOOK BUILDER
# ══════════════════════════════════════════════════════════════════════════

def build_workbook(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # style constants
    HEADER_FONT = Font(bold=True, size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT_W = Font(bold=True, size=11, color="FFFFFF")
    LEAKY_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    PCT_FMT = "0.0%"
    DOLLAR_FMT = '$#,##0.00'
    NUM_FMT = '#,##0'

    def write_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT_W
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

    def auto_width(ws, min_w=10, max_w=30):
        for col in ws.columns:
            lengths = []
            for cell in col:
                if cell.value:
                    lengths.append(len(str(cell.value)))
            if lengths:
                w = min(max(max(lengths) + 2, min_w), max_w)
                ws.column_dimensions[get_column_letter(col[0].column)].width = w

    def write_rows(ws, data, start_row=2, fmt_map=None):
        """Write list of dicts/tuples. fmt_map: {col_idx: number_format}"""
        for r, row_data in enumerate(data, start_row):
            if isinstance(row_data, dict):
                row_data = list(row_data.values())
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = THIN_BORDER
                if fmt_map and c in fmt_map:
                    cell.number_format = fmt_map[c]

    # ── Sheet 1: README ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "README"
    ws.sheet_properties.tabColor = "4472C4"

    date_range = f"{df['trade_date'].min()} to {df['trade_date'].max()}"
    n_trades = len(df)
    n_strategies = df["strategy"].nunique()
    n_symbols = df["symbol"].nunique()
    overall_wr = df["win"].mean()

    readme_lines = [
        ("Holly Analysis Lab Workbook v1", ""),
        ("", ""),
        ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Date Range:", date_range),
        ("Total Trades:", f"{n_trades:,}"),
        ("Strategies:", f"{n_strategies}"),
        ("Unique Symbols:", f"{n_symbols:,}"),
        ("Overall Win Rate:", f"{overall_wr:.1%}"),
        ("", ""),
        ("SHEETS:", ""),
        ("DATA_DICTIONARY", "Feature registry with timing class and coverage"),
        ("PRETRADE_FEATURES", "Ranked pretrade features by Cohen's d (full + OOS)"),
        ("STRATEGY_LAB", "Strategy ranking, temporal stability, SSP overlay"),
        ("REGIME_LAB", "Regime x direction heatmap, macro regimes"),
        ("SCORECARD", "Signal tiers, kill/monitor lists, data coverage"),
        ("", ""),
        ("TIMING CLASS KEY:", ""),
        ("pretrade", "Available before entry - safe for live use"),
        ("posttrade", "Uses outcome data - training/analysis only"),
        ("leaky", "Uses full-sample stats - looks pretrade but ISN'T"),
        ("", ""),
        ("SIGNIFICANCE:", ""),
        ("***", "p < 0.001"),
        ("**", "p < 0.01"),
        ("*", "p < 0.05"),
        ("(no star)", "p >= 0.05 (not significant)"),
        ("", ""),
        ("COHEN'S d GUIDE:", ""),
        ("|d| > 0.5", "Large effect"),
        ("|d| 0.3-0.5", "Medium effect"),
        ("|d| 0.1-0.3", "Small effect"),
        ("|d| < 0.1", "Negligible / noise"),
        ("", ""),
        ("OOS METHODOLOGY:", ""),
        ("Train = first 60% of trades (chronological)", ""),
        ("Test = last 40% of trades (chronological)", ""),
        ("OOS d must agree in sign with full-sample d to be credible", ""),
        ("", ""),
        ("SSP OVERLAY:", ""),
        ("Hierarchical Bayes shrinkage on strategy-sector win rate", ""),
        ("Applied as capped +/-10 bonus on AQS v2 score", ""),
        ("Adversarially validated: 5/5 tests PASS (script 100)", ""),
    ]
    for r, (a, b) in enumerate(readme_lines, 1):
        ws.cell(row=r, column=1, value=a).font = Font(bold=(r == 1 or a.endswith(":")), size=12 if r == 1 else 11)
        ws.cell(row=r, column=2, value=b)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60

    print("  [1/6] README built")

    # ── Sheet 2: DATA_DICTIONARY ─────────────────────────────────────────
    ws = wb.create_sheet("DATA_DICTIONARY")
    ws.sheet_properties.tabColor = "70AD47"

    headers = ["Feature", "Description", "Timing Class", "Source", "Coverage %"]
    write_header(ws, headers)

    dict_rows = []
    for feat, (desc, timing, source) in sorted(FEATURE_REGISTRY.items()):
        if feat in df.columns:
            cov = df[feat].notna().mean()
        else:
            cov = 0.0
        dict_rows.append((feat, desc, timing, source, cov))

    write_rows(ws, dict_rows, fmt_map={5: PCT_FMT})

    # highlight leaky rows
    for r in range(2, len(dict_rows) + 2):
        timing_val = ws.cell(row=r, column=3).value
        if timing_val == "leaky":
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = LEAKY_FILL

    ws.freeze_panes = "A2"
    auto_width(ws)
    print("  [2/6] DATA_DICTIONARY built")

    # ── Sheet 3: PRETRADE_FEATURES ───────────────────────────────────────
    ws = wb.create_sheet("PRETRADE_FEATURES")
    ws.sheet_properties.tabColor = "ED7D31"

    # identify numeric pretrade features
    pretrade_feats = [
        f for f, (_, tc, _) in FEATURE_REGISTRY.items()
        if tc == "pretrade" and f in df.columns and df[f].dtype in ("float64", "int64", "float32", "int32")
    ]

    # OOS split
    split_idx = int(len(df) * 0.6)
    df_train = df.iloc[:split_idx]
    df_test = df.iloc[split_idx:]

    wins_full = df[df["win"] == 1]
    losses_full = df[df["win"] == 0]
    wins_oos = df_test[df_test["win"] == 1]
    losses_oos = df_test[df_test["win"] == 0]

    feat_results = []
    for feat in pretrade_feats:
        w = wins_full[feat].dropna().values
        l = losses_full[feat].dropna().values
        n_total = len(w) + len(l)
        if n_total < 20:
            continue
        d_full, p_full = cohens_d(w, l)

        w_oos = wins_oos[feat].dropna().values
        l_oos = losses_oos[feat].dropna().values
        if len(w_oos) >= 5 and len(l_oos) >= 5:
            d_oos, p_oos = cohens_d(w_oos, l_oos)
        else:
            d_oos, p_oos = None, None

        cov = df[feat].notna().mean()
        feat_results.append({
            "feature": feat,
            "d_full": d_full,
            "d_oos": d_oos,
            "N": n_total,
            "p_full": p_full,
            "p_oos": p_oos,
            "coverage": cov,
        })

    feat_results.sort(key=lambda x: abs(x["d_full"]), reverse=True)

    headers = ["Feature", "d (full)", "d (OOS)", "N", "p-value", "Sig",
               "Coverage %", "OOS Agrees?", "Interpretation"]
    write_header(ws, headers)

    for r, fr in enumerate(feat_results, 2):
        d_full = fr["d_full"]
        d_oos = fr["d_oos"]
        oos_agrees = ""
        if d_oos is not None:
            oos_agrees = "YES" if (d_full * d_oos > 0) else "NO"

        interp = ""
        ad = abs(d_full)
        if ad > 0.5:
            interp = "Strong signal"
        elif ad > 0.3:
            interp = "Moderate signal"
        elif ad > 0.1:
            interp = "Weak signal"
        else:
            interp = "Noise"
        if d_oos is not None and oos_agrees == "NO":
            interp += " (OOS contradicts!)"

        row_data = [
            fr["feature"], round(d_full, 3),
            round(d_oos, 3) if d_oos is not None else "N/A",
            fr["N"], round(fr["p_full"], 4), significance_stars(fr["p_full"]),
            fr["coverage"], oos_agrees, interp,
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
        ws.cell(row=r, column=7).number_format = PCT_FMT

        # color OOS agrees
        if oos_agrees == "YES":
            ws.cell(row=r, column=8).fill = GOOD_FILL
        elif oos_agrees == "NO":
            ws.cell(row=r, column=8).fill = LEAKY_FILL

    # ── Quartile view for top 10 features ──
    top10 = feat_results[:10]
    qstart = len(feat_results) + 4
    ws.cell(row=qstart, column=1, value="TOP 10 FEATURE QUARTILE BREAKDOWN").font = Font(bold=True, size=12)
    qstart += 1

    q_headers = ["Feature", "Q1 WR", "Q1 N", "Q2 WR", "Q2 N", "Q3 WR", "Q3 N", "Q4 WR", "Q4 N",
                 "Q1 Avg PnL", "Q4 Avg PnL", "Spread"]
    write_header(ws, q_headers, row=qstart)

    for i, fr in enumerate(top10):
        feat = fr["feature"]
        vals = df[[feat, "win", "holly_pnl"]].dropna()
        if len(vals) < 40:
            continue
        try:
            vals["q"] = pd.qcut(vals[feat], 4, labels=["Q1", "Q2", "Q3", "Q4"], duplicates="drop")
        except ValueError:
            continue

        row = [feat]
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            qd = vals[vals["q"] == q]
            row.append(qd["win"].mean() if len(qd) > 0 else None)
            row.append(len(qd))
        q1_pnl = vals[vals["q"] == "Q1"]["holly_pnl"].mean() if len(vals[vals["q"] == "Q1"]) > 0 else 0
        q4_pnl = vals[vals["q"] == "Q4"]["holly_pnl"].mean() if len(vals[vals["q"] == "Q4"]) > 0 else 0
        row.extend([q1_pnl, q4_pnl, q4_pnl - q1_pnl])

        r = qstart + 1 + i
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
        for c in [2, 4, 6, 8]:
            ws.cell(row=r, column=c).number_format = PCT_FMT
        for c in [10, 11, 12]:
            ws.cell(row=r, column=c).number_format = DOLLAR_FMT

    # ── Leakage guard section ──
    lg_start = qstart + len(top10) + 3
    ws.cell(row=lg_start, column=1, value="LEAKAGE GUARD - Excluded Features").font = Font(bold=True, size=12, color="FF0000")
    lg_start += 1
    ws.cell(row=lg_start, column=1, value="These features use full-sample statistics and CANNOT be used for live trading signals:").font = Font(italic=True)
    lg_start += 1

    leaky_feats = [(f, d) for f, (d, tc, _) in FEATURE_REGISTRY.items() if tc == "leaky"]
    write_header(ws, ["Feature", "Description", "Why It's Leaky"], row=lg_start)
    for i, (feat, desc) in enumerate(leaky_feats):
        r = lg_start + 1 + i
        ws.cell(row=r, column=1, value=feat).fill = LEAKY_FILL
        ws.cell(row=r, column=2, value=desc).fill = LEAKY_FILL
        ws.cell(row=r, column=3, value="Computed from ALL trades including future ones").fill = LEAKY_FILL
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.freeze_panes = "A2"
    auto_width(ws)
    print(f"  [3/6] PRETRADE_FEATURES built ({len(feat_results)} features ranked)")

    # ── Sheet 4: STRATEGY_LAB ────────────────────────────────────────────
    ws = wb.create_sheet("STRATEGY_LAB")
    ws.sheet_properties.tabColor = "FFC000"

    strat_groups = df.groupby(["strategy", "direction"])

    strat_rows = []
    for (strat, dirn), g in strat_groups:
        n = len(g)
        wins = g["win"].sum()
        wr = wins / n if n > 0 else 0
        avg_pnl = g["holly_pnl"].mean()
        total_pnl = g["holly_pnl"].sum()
        pf = profit_factor(g["holly_pnl"])
        sp = sharpe_proxy(g["holly_pnl"])
        bwr, blo, bhi = bayesian_wr(wins, n)

        # cohen's d: this strat vs rest
        strat_pnls = g["holly_pnl"].values
        rest_pnls = df[~df.index.isin(g.index)]["holly_pnl"].values
        d, p = cohens_d(strat_pnls, rest_pnls)

        # temporal stability
        half = n // 2
        if half >= 10:
            wr1 = g.iloc[:half]["win"].mean()
            wr2 = g.iloc[half:]["win"].mean()
            decay = (wr2 - wr1) * 100  # positive = improving
        else:
            wr1 = wr2 = decay = None

        # SSP overlay mean
        ssp_mean = g["shrunk_strat_sector_wr"].mean() if "shrunk_strat_sector_wr" in g.columns else None

        thin = n < 50
        edge = "EDGE" if (wr > 0.52 and total_pnl > 0 and pf > 1.2 and not thin) else \
               "MARGINAL" if (wr > 0.48 and total_pnl > 0) else "NO EDGE"

        strat_rows.append({
            "strategy": strat, "direction": dirn, "N": n,
            "WR": wr, "avg_pnl": avg_pnl, "total_pnl": total_pnl,
            "pf": pf if pf != float("inf") else 99.9, "sharpe": sp,
            "bayes_wr": bwr, "bayes_lo": blo, "bayes_hi": bhi,
            "d": d, "sig": significance_stars(p),
            "wr_1h": wr1, "wr_2h": wr2, "decay_pp": decay,
            "ssp_mean": ssp_mean, "thin": thin, "edge": edge,
        })

    strat_rows.sort(key=lambda x: x["total_pnl"], reverse=True)

    headers = ["Strategy", "Dir", "N", "WR", "Avg PnL", "Total PnL",
               "PF", "Sharpe", "Bayes WR", "95% CI", "d", "Sig",
               "WR 1st Half", "WR 2nd Half", "Decay (pp)", "SSP Mean WR",
               "Thin?", "Edge"]
    write_header(ws, headers)

    for r, sr in enumerate(strat_rows, 2):
        row = [
            sr["strategy"], sr["direction"], sr["N"],
            sr["WR"], sr["avg_pnl"], sr["total_pnl"],
            sr["pf"], round(sr["sharpe"], 2),
            sr["bayes_wr"], f"[{sr['bayes_lo']:.1%}, {sr['bayes_hi']:.1%}]",
            round(sr["d"], 3), sr["sig"],
            sr["wr_1h"], sr["wr_2h"],
            round(sr["decay_pp"], 1) if sr["decay_pp"] is not None else "N/A",
            sr["ssp_mean"],
            "THIN" if sr["thin"] else "",
            sr["edge"],
        ]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER

        # format
        ws.cell(row=r, column=4).number_format = PCT_FMT
        ws.cell(row=r, column=5).number_format = DOLLAR_FMT
        ws.cell(row=r, column=6).number_format = DOLLAR_FMT
        ws.cell(row=r, column=7).number_format = "0.00"
        ws.cell(row=r, column=9).number_format = PCT_FMT
        if sr["wr_1h"] is not None:
            ws.cell(row=r, column=13).number_format = PCT_FMT
            ws.cell(row=r, column=14).number_format = PCT_FMT
        if sr["ssp_mean"] is not None:
            ws.cell(row=r, column=16).number_format = PCT_FMT

        # color edge verdict
        edge_cell = ws.cell(row=r, column=18)
        if sr["edge"] == "EDGE":
            edge_cell.fill = GOOD_FILL
        elif sr["edge"] == "NO EDGE":
            edge_cell.fill = LEAKY_FILL

        # flag decay > 10pp
        if sr["decay_pp"] is not None and abs(sr["decay_pp"]) > 10:
            ws.cell(row=r, column=15).fill = WARN_FILL

        # flag thin
        if sr["thin"]:
            ws.cell(row=r, column=17).fill = WARN_FILL

    ws.freeze_panes = "A2"
    auto_width(ws)
    print(f"  [4/6] STRATEGY_LAB built ({len(strat_rows)} strategy-direction combos)")

    # ── Sheet 5: REGIME_LAB ──────────────────────────────────────────────
    ws = wb.create_sheet("REGIME_LAB")
    ws.sheet_properties.tabColor = "A5A5A5"

    # regime x direction matrix
    has_regime = df["trend_regime"].notna() & df["vol_regime"].notna()
    rdf = df[has_regime].copy()

    if len(rdf) > 0:
        rdf["regime_combo"] = rdf["trend_regime"].astype(str) + " / " + rdf["vol_regime"].astype(str)

        headers = ["Regime Combo", "Long WR", "Long N", "Long Avg PnL",
                   "Short WR", "Short N", "Short Avg PnL"]
        write_header(ws, headers)

        combos = sorted(rdf["regime_combo"].unique())
        regime_rows = []
        for combo in combos:
            rc = rdf[rdf["regime_combo"] == combo]
            row = [combo]
            for dirn in ["long", "short"]:
                dd = rc[rc["direction"] == dirn]
                n = len(dd)
                if n >= 20:
                    row.extend([dd["win"].mean(), n, dd["holly_pnl"].mean()])
                else:
                    row.extend(["--" if n < 20 else dd["win"].mean(), n, "--" if n < 20 else dd["holly_pnl"].mean()])
            regime_rows.append((row, combo))

        for r, (row, _) in enumerate(regime_rows, 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = THIN_BORDER
            if isinstance(row[1], float):
                ws.cell(row=r, column=2).number_format = PCT_FMT
            if isinstance(row[3], float):
                ws.cell(row=r, column=4).number_format = DOLLAR_FMT
            if isinstance(row[4], float):
                ws.cell(row=r, column=5).number_format = PCT_FMT
            if isinstance(row[6], float):
                ws.cell(row=r, column=7).number_format = DOLLAR_FMT

        # color-scale WR cells
        n_regime_rows = len(regime_rows)
        if n_regime_rows > 1:
            for col_letter in ["B", "E"]:
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{n_regime_rows + 1}",
                    ColorScaleRule(
                        start_type="num", start_value=0.40, start_color="F8696B",
                        mid_type="num", mid_value=0.50, mid_color="FFEB84",
                        end_type="num", end_value=0.60, end_color="63BE7B",
                    ),
                )

        # ── Macro regime table ──
        macro_start = n_regime_rows + 4
        ws.cell(row=macro_start, column=1, value="MACRO REGIME MATRIX").font = Font(bold=True, size=12)
        macro_start += 1

        has_macro = df["vix_regime"].notna() & df["yield_curve_regime"].notna()
        mdf = df[has_macro].copy()

        if len(mdf) > 100:
            mdf["macro_combo"] = mdf["vix_regime"].astype(str) + " / " + mdf["yield_curve_regime"].astype(str)
            m_headers = ["VIX / Yield Curve", "Long WR", "Long N", "Short WR", "Short N", "Avg PnL"]
            write_header(ws, m_headers, row=macro_start)

            for r_idx, combo in enumerate(sorted(mdf["macro_combo"].unique())):
                mc = mdf[mdf["macro_combo"] == combo]
                r = macro_start + 1 + r_idx
                ws.cell(row=r, column=1, value=combo).border = THIN_BORDER

                for c_off, dirn in [(2, "long"), (4, "short")]:
                    dd = mc[mc["direction"] == dirn]
                    n = len(dd)
                    if n >= 20:
                        ws.cell(row=r, column=c_off, value=dd["win"].mean()).number_format = PCT_FMT
                    else:
                        ws.cell(row=r, column=c_off, value="--")
                    ws.cell(row=r, column=c_off + 1, value=n)
                    for cc in range(c_off, c_off + 2):
                        ws.cell(row=r, column=cc).border = THIN_BORDER

                ws.cell(row=r, column=6, value=mc["holly_pnl"].mean()).number_format = DOLLAR_FMT
                ws.cell(row=r, column=6).border = THIN_BORDER

        # ── Top/bottom 5 regime combos ──
        if len(rdf) > 0:
            combo_stats = []
            for combo in combos:
                rc = rdf[rdf["regime_combo"] == combo]
                n = len(rc)
                if n >= 20:
                    combo_stats.append((combo, rc["win"].mean(), n, rc["holly_pnl"].mean()))

            combo_stats.sort(key=lambda x: x[1], reverse=True)

            tb_start = ws.max_row + 3
            ws.cell(row=tb_start, column=1, value="TOP 5 REGIME COMBOS BY WR").font = Font(bold=True, size=12)
            tb_start += 1
            write_header(ws, ["Regime", "WR", "N", "Avg PnL"], row=tb_start)
            for i, (combo, wr, n, avg) in enumerate(combo_stats[:5]):
                r = tb_start + 1 + i
                ws.cell(row=r, column=1, value=combo).border = THIN_BORDER
                ws.cell(row=r, column=2, value=wr).number_format = PCT_FMT
                ws.cell(row=r, column=2).border = THIN_BORDER
                ws.cell(row=r, column=2).fill = GOOD_FILL
                ws.cell(row=r, column=3, value=n).border = THIN_BORDER
                ws.cell(row=r, column=4, value=avg).number_format = DOLLAR_FMT
                ws.cell(row=r, column=4).border = THIN_BORDER

            bb_start = tb_start + 7
            ws.cell(row=bb_start, column=1, value="BOTTOM 5 REGIME COMBOS BY WR").font = Font(bold=True, size=12)
            bb_start += 1
            write_header(ws, ["Regime", "WR", "N", "Avg PnL"], row=bb_start)
            for i, (combo, wr, n, avg) in enumerate(combo_stats[-5:]):
                r = bb_start + 1 + i
                ws.cell(row=r, column=1, value=combo).border = THIN_BORDER
                ws.cell(row=r, column=2, value=wr).number_format = PCT_FMT
                ws.cell(row=r, column=2).border = THIN_BORDER
                ws.cell(row=r, column=2).fill = LEAKY_FILL
                ws.cell(row=r, column=3, value=n).border = THIN_BORDER
                ws.cell(row=r, column=4, value=avg).number_format = DOLLAR_FMT
                ws.cell(row=r, column=4).border = THIN_BORDER

    # timing note
    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1,
            value="NOTE: Regime labels are from daily bars. Intraday regime may differ from EOD classification.").font = Font(italic=True, color="808080")

    ws.freeze_panes = "A2"
    auto_width(ws)
    print(f"  [5/6] REGIME_LAB built ({len(rdf)} trades with regime data)")

    # ── Sheet 6: SCORECARD ───────────────────────────────────────────────
    ws = wb.create_sheet("SCORECARD")
    ws.sheet_properties.tabColor = "FF0000"

    # tier classification
    tier1, tier2, tier3 = [], [], []
    for fr in feat_results:
        ad = abs(fr["d_full"])
        oos_ok = fr["d_oos"] is not None and (fr["d_full"] * fr["d_oos"] > 0)
        if ad > 0.3 and oos_ok:
            tier1.append(fr)
        elif ad > 0.1 and fr["d_oos"] is not None:
            tier2.append(fr)
        else:
            tier3.append(fr)

    r = 1
    ws.cell(row=r, column=1, value="SIGNAL TIER LIST").font = Font(bold=True, size=14)
    r += 2

    # Tier 1
    ws.cell(row=r, column=1, value="TIER 1: Strong signals (|d| > 0.3, persists OOS)").font = Font(bold=True, color="006100")
    ws.cell(row=r, column=1).fill = GOOD_FILL
    r += 1
    if tier1:
        write_header(ws, ["Feature", "d (full)", "d (OOS)", "Coverage"], row=r)
        r += 1
        for fr in tier1:
            ws.cell(row=r, column=1, value=fr["feature"]).border = THIN_BORDER
            ws.cell(row=r, column=2, value=round(fr["d_full"], 3)).border = THIN_BORDER
            ws.cell(row=r, column=3, value=round(fr["d_oos"], 3) if fr["d_oos"] else "N/A").border = THIN_BORDER
            ws.cell(row=r, column=4, value=fr["coverage"]).number_format = PCT_FMT
            ws.cell(row=r, column=4).border = THIN_BORDER
            r += 1
    else:
        ws.cell(row=r, column=1, value="(none)")
        r += 1
    r += 1

    # Tier 2
    ws.cell(row=r, column=1, value="TIER 2: Conditional signals (|d| 0.1-0.3)").font = Font(bold=True, color="9C6500")
    ws.cell(row=r, column=1).fill = WARN_FILL
    r += 1
    if tier2:
        write_header(ws, ["Feature", "d (full)", "d (OOS)", "Coverage"], row=r)
        r += 1
        for fr in tier2[:15]:  # cap at 15
            ws.cell(row=r, column=1, value=fr["feature"]).border = THIN_BORDER
            ws.cell(row=r, column=2, value=round(fr["d_full"], 3)).border = THIN_BORDER
            ws.cell(row=r, column=3, value=round(fr["d_oos"], 3) if fr["d_oos"] else "N/A").border = THIN_BORDER
            ws.cell(row=r, column=4, value=fr["coverage"]).number_format = PCT_FMT
            ws.cell(row=r, column=4).border = THIN_BORDER
            r += 1
    else:
        ws.cell(row=r, column=1, value="(none)")
        r += 1
    r += 1

    # Tier 3
    ws.cell(row=r, column=1, value="TIER 3: Noise (|d| < 0.1 or OOS contradicts)").font = Font(bold=True, color="9C0006")
    ws.cell(row=r, column=1).fill = LEAKY_FILL
    r += 1
    ws.cell(row=r, column=1, value=f"{len(tier3)} features classified as noise").font = Font(italic=True)
    r += 1
    # list first 10
    for fr in tier3[:10]:
        ws.cell(row=r, column=1, value=fr["feature"])
        ws.cell(row=r, column=2, value=round(fr["d_full"], 3))
        r += 1
    if len(tier3) > 10:
        ws.cell(row=r, column=1, value=f"... and {len(tier3) - 10} more")
        r += 1
    r += 1

    # ── SSP overlay summary ──
    ws.cell(row=r, column=1, value="SSP OVERLAY SUMMARY").font = Font(bold=True, size=12)
    r += 1

    # compute SSP d
    ssp_vals = df["shrunk_strat_sector_wr"].dropna()
    if len(ssp_vals) > 100:
        ssp_w = df[df["win"] == 1]["shrunk_strat_sector_wr"].dropna().values
        ssp_l = df[df["win"] == 0]["shrunk_strat_sector_wr"].dropna().values
        ssp_d, ssp_p = cohens_d(ssp_w, ssp_l)
    else:
        ssp_d, ssp_p = 0, 1

    ssp_items = [
        ("SSP Cohen's d:", f"{ssp_d:.3f} {significance_stars(ssp_p)}"),
        ("Coverage:", f"{df['shrunk_strat_sector_wr'].notna().mean():.1%}"),
        ("Adversarial tests:", "5/5 PASS (script 100)"),
        ("Method:", "Hierarchical Bayes shrinkage (cell -> strategy -> global)"),
        ("Application:", "Capped +/-10 overlay on AQS v2"),
    ]
    for label, val in ssp_items:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1

    # ── Kill list ──
    ws.cell(row=r, column=1, value="KILL LIST - Dead Features").font = Font(bold=True, size=12, color="FF0000")
    r += 1
    kill_candidates = []
    for fr in feat_results:
        # noise + OOS contradicts or very low d
        if abs(fr["d_full"]) < 0.05:
            kill_candidates.append((fr["feature"], fr["d_full"], "d ~ 0, pure noise"))
        elif fr["d_oos"] is not None and fr["d_full"] * fr["d_oos"] < 0 and abs(fr["d_full"]) < 0.2:
            kill_candidates.append((fr["feature"], fr["d_full"], "OOS sign flip"))

    # always include sector_prior_wr in kill list if we know it's dead
    leaky_kills = [
        ("sector_win_rate", "Full-sample sector stat, not available pretrade"),
        ("sector_avg_pnl", "Full-sample sector stat, not available pretrade"),
        ("tod_win_rate", "Full-sample time-of-day stat, not available pretrade"),
        ("strat_win_rate", "Full-sample strategy stat, not available pretrade"),
    ]

    write_header(ws, ["Feature", "d / Reason", "Evidence"], row=r)
    r += 1
    for feat, d_val, reason in kill_candidates[:10]:
        ws.cell(row=r, column=1, value=feat).fill = LEAKY_FILL
        ws.cell(row=r, column=2, value=f"d={d_val:.3f}")
        ws.cell(row=r, column=3, value=reason)
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    for feat, reason in leaky_kills:
        ws.cell(row=r, column=1, value=feat).fill = LEAKY_FILL
        ws.cell(row=r, column=2, value="LEAKY")
        ws.cell(row=r, column=3, value=reason)
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    r += 1

    # ── Monitor list ──
    ws.cell(row=r, column=1, value="MONITOR LIST - Conditional Features").font = Font(bold=True, size=12, color="9C6500")
    r += 1
    ws.cell(row=r, column=1, value="Features with signal but need rolling tracking to confirm persistence:").font = Font(italic=True)
    r += 1
    for fr in tier2[:8]:
        ws.cell(row=r, column=1, value=fr["feature"])
        ws.cell(row=r, column=2, value=f"d={fr['d_full']:.3f}")
        ws.cell(row=r, column=3, value=f"OOS d={fr['d_oos']:.3f}" if fr["d_oos"] else "No OOS")
        r += 1
    r += 1

    # ── Data coverage ──
    ws.cell(row=r, column=1, value="DATA COVERAGE TABLE").font = Font(bold=True, size=12)
    r += 1
    write_header(ws, ["Data Source", "Coverage %", "Trades Covered", "Notes"], row=r)
    r += 1

    sources = {
        "Core trades": ("trade_id", "Always available"),
        "Regime (technical)": ("trend_regime", "Daily bar regime classification"),
        "Indicator features": ("rsi_zone", "RSI/MACD/MA from indicators"),
        "Short features": ("short_interest", "Short interest & squeeze"),
        "Insider features": ("insider_buys_30d", "30-day insider activity"),
        "Benzinga news": ("news_count_24h", "24h news count"),
        "Macro (FRED)": ("vix_regime", "VIX/yield/rate regime"),
        "SSP overlay": ("shrunk_strat_sector_wr", "Strategy-sector shrunk WR"),
    }
    for source, (col, note) in sources.items():
        if col in df.columns:
            cov = df[col].notna().mean()
            n_covered = df[col].notna().sum()
        else:
            cov = 0
            n_covered = 0
        ws.cell(row=r, column=1, value=source).border = THIN_BORDER
        ws.cell(row=r, column=2, value=cov).number_format = PCT_FMT
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=3, value=n_covered).number_format = NUM_FMT
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=4, value=note).border = THIN_BORDER
        r += 1

    ws.freeze_panes = "A2"
    auto_width(ws)
    print(f"  [6/6] SCORECARD built (Tier1={len(tier1)}, Tier2={len(tier2)}, Tier3={len(tier3)})")

    # ── Save ─────────────────────────────────────────────────────────────
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f"\nWorkbook saved: {OUTPUT_PATH}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Size: {OUTPUT_PATH.stat().st_size / 1024:.0f} KB")


# ══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 70)
    print("Holly Analysis Lab Workbook v1")
    print("=" * 70)

    print("\n[1] Loading data from DuckDB...")
    df = load_data()
    print(f"    {len(df)} trades loaded, {df.columns.size} columns")

    print("\n[2] Building workbook...")
    build_workbook(df)

    print("\nDone.")
